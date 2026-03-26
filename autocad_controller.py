"""
AutoCAD控制核心模块
使用pywin32和COM接口与AutoCAD进行通信
"""

import win32com.client
import pythoncom
import time
import math
from typing import Optional, List
import logging

try:
    import win32gui
    import win32api
    import win32con
except Exception:
    win32gui = None
    win32api = None
    win32con = None

logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s'
)
logger = logging.getLogger(__name__)


def _format_com_error(e: Exception) -> str:
    """格式化 COM 异常信息，避免 <unknown>"""
    msg = str(e).strip()
    if msg and msg != "<unknown>":
        return msg
    try:
        if hasattr(e, "args") and e.args:
            return repr(e.args)
        return repr(e)
    except Exception:
        return "COM调用异常"


def _format_geometry_key(key: str) -> str:
    """将几何属性键转换为友好的中文显示名"""
    key_map = {
        "center_x": "圆心X",
        "center_y": "圆心Y", 
        "center_z": "圆心Z",
        "start_x": "起点X",
        "start_y": "起点Y",
        "start_z": "起点Z",
        "end_x": "终点X",
        "end_y": "终点Y",
        "end_z": "终点Z",
        "position_x": "位置X",
        "position_y": "位置Y",
        "position_z": "位置Z",
        "insert_x": "插入点X",
        "insert_y": "插入点Y",
        "insert_z": "插入点Z",
        "x": "X坐标",
        "y": "Y坐标",
        "z": "Z坐标",
        "radius": "半径",
        "diameter": "直径",
        "length": "长度",
        "width": "宽度",
        "height": "高度/字高",
        "area": "面积",
        "circumference": "周长",
        "arc_length": "弧长",
        "angle": "角度",
        "rotation": "旋转角度",
        "start_angle_deg": "起始角度(度)",
        "end_angle_deg": "终止角度(度)",
        "start_angle_rad": "起始角度(弧度)",
        "end_angle_rad": "终止角度(弧度)",
        "major_radius": "长轴半径",
        "minor_radius": "短轴半径",
        "vertices_count": "顶点数量",
        "vertices_list": "顶点坐标列表",
        "closed": "是否闭合",
        "degree": "阶数",
        "control_points_count": "控制点数",
        "text": "文本内容",
        "pattern": "填充图案",
        "block_name": "块名称",
        "measurement": "测量值",
        "color": "颜色",
        "linetype": "线型",
        "lineweight": "线宽",
        "draw_command": "绘制命令",
        "scale_x": "X缩放",
        "scale_y": "Y缩放",
        "scale_z": "Z缩放"
    }
    return key_map.get(key, key)


class AutoCADController:
    """AutoCAD控制器类"""
    
    def __init__(self):
        self.acad_app = None
        self.acad_doc = None
        self.is_connected = False
        self.connection_timeout = 10

    def ensure_document(self) -> bool:
        """重新获取当前活动文档，避免引用失效（切换文档、停止后再次会话等）"""
        if not self.acad_app:
            logger.warning("ensure_document: acad_app 为空")
            return False
        try:
            # 每次都重新获取活动文档，确保引用有效
            self.acad_doc = self.acad_app.ActiveDocument
            if self.acad_doc is None:
                logger.warning("ensure_document: ActiveDocument 返回 None，请先打开 DWG 图纸")
                return False
            # 额外验证：尝试访问文档属性
            _ = self.acad_doc.Name
            return True
        except Exception as e:
            logger.warning(f"ensure_document 失败: {_format_com_error(e)}")
            self.acad_doc = None
            return False

    def connect(self, timeout: int = 10) -> bool:
        """连接到AutoCAD实例"""
        try:
            logger.info("正在尝试连接到AutoCAD...")
            
            pythoncom.CoInitialize()
            
            try:
                self.acad_app = win32com.client.GetActiveObject("AutoCAD.Application")
                logger.info("成功连接到现有的AutoCAD实例")
            except Exception as e:
                logger.warning(f"未找到运行中的AutoCAD实例: {e}")
                logger.info("请先启动AutoCAD，然后点击'连接AutoCAD'按钮")
                pythoncom.CoUninitialize()
                return False
            
            try:
                self.acad_doc = self.acad_app.ActiveDocument
            except Exception as e:
                logger.warning(f"无法获取活动文档: {e}")
                self.acad_doc = None
            
            self.is_connected = True
            
            try:
                version = self.acad_app.Version
                logger.info(f"已连接到AutoCAD {version}")
            except Exception as e:
                logger.warning(f"无法获取版本信息: {e}")
            
            return True
            
        except Exception as e:
            logger.error(f"连接AutoCAD失败: {e}")
            self.is_connected = False
            return False
    
    def disconnect(self):
        """断开与AutoCAD的连接"""
        try:
            self.acad_doc = None
            self.acad_app = None
            self.is_connected = False
            logger.info("已断开与AutoCAD的连接")
            pythoncom.CoUninitialize()
        except Exception as e:
            logger.error(f"断开连接时出错: {e}")
    
    def send_command(self, command: str, delay: float = 0.5) -> bool:
        """发送命令到AutoCAD。delay=0 时不等待，便于队列模式下主线程不阻塞、停止可点。"""
        if not self.is_connected or not self.acad_app:
            logger.error("未连接到AutoCAD")
            return False

        # 每次发送前刷新活动文档，避免“停止后再次会话”或切换文档后引用失效
        if not self.ensure_document():
            logger.error("无法获取活动文档，请确认 AutoCAD 已打开图纸")
            return False

        if not command.endswith('\n'):
            command = command + '\n'

        def do_send():
            try:
                self.acad_doc.SendCommand(command)
                return True
            except Exception as e1:
                logger.warning(f"SendCommand失败: {_format_com_error(e1)}，尝试 SendStringToExecute")
                try:
                    self.acad_doc.SendStringToExecute(command, True, False, False)
                    return True
                except Exception as e2:
                    logger.error(f"SendStringToExecute也失败: {_format_com_error(e2)}")
                    raise e2

        try:
            logger.info(f"发送命令: {command}")
            do_send()
            if delay > 0:
                time.sleep(delay)
            return True
        except Exception as e:
            logger.error(f"发送命令失败: {_format_com_error(e)}")
            # 失败时尝试刷新文档并重试一次
            try:
                if self.ensure_document():
                    logger.info("已刷新活动文档，正在重试发送命令...")
                    do_send()
                    if delay > 0:
                        time.sleep(delay)
                    return True
            except Exception as retry_e:
                logger.error(f"重试发送命令失败: {_format_com_error(retry_e)}")
            return False
    
    def send_commands(self, commands: List[str], delay: float = 0.5) -> bool:
        """批量发送命令"""
        success = True
        for cmd in commands:
            if not self.send_command(cmd, delay):
                success = False
        return success
    
    def _activate_acad_window(self) -> bool:
        """尝试激活 AutoCAD 窗口。仅在需要键盘兜底时使用。"""
        if not self.acad_app:
            return False
        if win32gui is None:
            return False
        try:
            hwnd = int(self.acad_app.HWND)
            if hwnd:
                try:
                    win32gui.ShowWindow(hwnd, 9)  # SW_RESTORE
                except Exception:
                    pass
                win32gui.SetForegroundWindow(hwnd)
                time.sleep(0.05)
                return True
        except Exception as e:
            logger.debug(f"激活AutoCAD窗口失败: {_format_com_error(e)}")
        return False

    def _send_keyboard_cancel_to_acad(self) -> bool:
        """键盘兜底：向 AutoCAD 窗口发送 Ctrl+C / ESC（不依赖本工具窗口焦点）。"""
        if win32api is None or win32con is None:
            return False
        if not self._activate_acad_window():
            return False

        try:
            # Ctrl down
            win32api.keybd_event(win32con.VK_CONTROL, 0, 0, 0)
            # C down/up
            win32api.keybd_event(ord('C'), 0, 0, 0)
            win32api.keybd_event(ord('C'), 0, win32con.KEYEVENTF_KEYUP, 0)
            # C 再一次
            win32api.keybd_event(ord('C'), 0, 0, 0)
            win32api.keybd_event(ord('C'), 0, win32con.KEYEVENTF_KEYUP, 0)
            # Ctrl up
            win32api.keybd_event(win32con.VK_CONTROL, 0, win32con.KEYEVENTF_KEYUP, 0)
            time.sleep(0.03)

            # ESC down/up
            win32api.keybd_event(win32con.VK_ESCAPE, 0, 0, 0)
            win32api.keybd_event(win32con.VK_ESCAPE, 0, win32con.KEYEVENTF_KEYUP, 0)
            time.sleep(0.03)

            logger.info("已发送键盘取消命令（Ctrl+C, Ctrl+C, ESC）")
            return True
        except Exception as e:
            logger.debug(f"发送键盘取消失败: {_format_com_error(e)}")
            return False

    def cancel_command(self) -> bool:
        """取消当前命令（COM 方式，不依赖窗口焦点）。返回是否发送成功。"""
        if not self.is_connected or not self.acad_app:
            return False

        try:
            doc = self.acad_app.ActiveDocument
        except Exception as e:
            logger.debug(f"取消命令时获取活动文档失败: {_format_com_error(e)}")
            return False

        cancel_sequences = [
            "\x03\x03\n",      # Ctrl+C Ctrl+C（硬取消）
            "\x03\n",           # 再补一次 Ctrl+C
            chr(27) + "\n",     # ESC
            "\n",               # 回车，清理残留提示
        ]

        sent_any = False
        for seq in cancel_sequences:
            try:
                doc.SendCommand(seq)
                sent_any = True
                time.sleep(0.06)
            except Exception as e:
                logger.debug(f"发送取消序列失败: {_format_com_error(e)}")

        if sent_any:
            logger.info("已发送取消命令（COM: ^C^C/ESC）")
        return sent_any

    def force_cancel_command(self, rounds: int = 8, interval: float = 0.08) -> bool:
        """强制取消：多轮 COM 取消；必要时使用键盘事件兜底。"""
        success = False

        # 第一阶段：COM 多轮取消
        for _ in range(max(1, rounds)):
            success = self.cancel_command() or success
            time.sleep(interval)

        # 第二阶段：激活 AutoCAD 并再次 COM 取消
        if self._activate_acad_window():
            for _ in range(3):
                success = self.cancel_command() or success
                time.sleep(interval)

        # 第三阶段：键盘级兜底（Ctrl+C/Ctrl+C/ESC），用于某些 COM 取消无效场景
        kb_success = False
        for _ in range(3):
            kb_success = self._send_keyboard_cancel_to_acad() or kb_success
            time.sleep(interval)

        return success or kb_success

    # ==================== COM API 直接绘图方法 ====================
    # 绕过命令行，直接使用 AutoCAD COM API 创建图形对象
    # 优势：完全自动化、无需用户干预、执行速度快、支持撤销

    def _make_point(self, coords: tuple) -> win32com.client.VARIANT:
        """创建 AutoCAD 3D 点坐标（COM VARIANT 数组）"""
        if len(coords) >= 3:
            x, y, z = float(coords[0]), float(coords[1]), float(coords[2])
        elif len(coords) == 2:
            x, y, z = float(coords[0]), float(coords[1]), 0.0
        else:
            x, y, z = 0.0, 0.0, 0.0
        return win32com.client.VARIANT(pythoncom.VT_ARRAY | pythoncom.VT_R8, [x, y, z])

    def _make_point_2d(self, coords: list) -> win32com.client.VARIANT:
        """创建 2D 点坐标数组（用于多段线等）"""
        flat = []
        for pt in coords:
            flat.append(float(pt[0]))
            flat.append(float(pt[1]))
        return win32com.client.VARIANT(pythoncom.VT_ARRAY | pythoncom.VT_R8, flat)

    def draw_line(self, start: tuple, end: tuple, layer: str = None) -> dict:
        """
        绘制直线
        
        Args:
            start: 起点 (x, y) 或 (x, y, z)
            end: 终点 (x, y) 或 (x, y, z)
            layer: 图层名（可选）
        
        Returns:
            {"success": bool, "message": str, "entity": object}
        """
        if not self.ensure_document():
            return {"success": False, "message": "未连接到AutoCAD或无活动文档"}
        
        try:
            start_point = self._make_point(start)
            end_point = self._make_point(end)
            
            line = self.acad_doc.ModelSpace.AddLine(start_point, end_point)
            
            if layer:
                try:
                    line.Layer = layer
                except Exception:
                    pass  # 图层不存在时忽略
            
            logger.info(f"已绘制直线: {start} -> {end}")
            return {"success": True, "message": f"直线已绘制", "entity": line}
            
        except Exception as e:
            err = _format_com_error(e)
            logger.error(f"绘制直线失败: {err}")
            return {"success": False, "message": f"绘制直线失败: {err}"}

    def draw_circle(self, center: tuple, radius: float, layer: str = None) -> dict:
        """
        绘制圆
        
        Args:
            center: 圆心 (x, y) 或 (x, y, z)
            radius: 半径
            layer: 图层名（可选）
        
        Returns:
            {"success": bool, "message": str, "entity": object}
        """
        if not self.ensure_document():
            return {"success": False, "message": "未连接到AutoCAD或无活动文档"}
        
        try:
            center_point = self._make_point(center)
            circle = self.acad_doc.ModelSpace.AddCircle(center_point, float(radius))
            
            if layer:
                try:
                    circle.Layer = layer
                except Exception:
                    pass
            
            logger.info(f"已绘制圆: 圆心{center}, 半径{radius}")
            return {"success": True, "message": f"圆已绘制（半径={radius}）", "entity": circle}
            
        except Exception as e:
            err = _format_com_error(e)
            logger.error(f"绘制圆失败: {err}")
            return {"success": False, "message": f"绘制圆失败: {err}"}

    def draw_rectangle(self, corner1: tuple, corner2: tuple, layer: str = None) -> dict:
        """
        绘制矩形（闭合多段线）
        
        Args:
            corner1: 角点1 (x, y)
            corner2: 对角点2 (x, y)
            layer: 图层名（可选）
        
        Returns:
            {"success": bool, "message": str, "entity": object}
        """
        if not self.ensure_document():
            return {"success": False, "message": "未连接到AutoCAD或无活动文档"}
        
        try:
            x1, y1 = float(corner1[0]), float(corner1[1])
            x2, y2 = float(corner2[0]), float(corner2[1])
            
            # 四个顶点（闭合）
            points = self._make_point_2d([
                (x1, y1), (x2, y1), (x2, y2), (x1, y2)
            ])
            
            pline = self.acad_doc.ModelSpace.AddLightweightPolyline(points)
            pline.Closed = True
            
            if layer:
                try:
                    pline.Layer = layer
                except Exception:
                    pass
            
            logger.info(f"已绘制矩形: ({x1},{y1}) - ({x2},{y2})")
            return {"success": True, "message": f"矩形已绘制", "entity": pline}
            
        except Exception as e:
            err = _format_com_error(e)
            logger.error(f"绘制矩形失败: {err}")
            return {"success": False, "message": f"绘制矩形失败: {err}"}

    def draw_arc(self, center: tuple, radius: float, start_angle: float, end_angle: float, layer: str = None) -> dict:
        """
        绘制圆弧
        
        Args:
            center: 圆心 (x, y) 或 (x, y, z)
            radius: 半径
            start_angle: 起始角度（弧度）
            end_angle: 结束角度（弧度）
            layer: 图层名（可选）
        
        Returns:
            {"success": bool, "message": str, "entity": object}
        """
        if not self.ensure_document():
            return {"success": False, "message": "未连接到AutoCAD或无活动文档"}
        
        try:
            center_point = self._make_point(center)
            arc = self.acad_doc.ModelSpace.AddArc(
                center_point, 
                float(radius), 
                float(start_angle), 
                float(end_angle)
            )
            
            if layer:
                try:
                    arc.Layer = layer
                except Exception:
                    pass
            
            logger.info(f"已绘制圆弧: 圆心{center}, 半径{radius}")
            return {"success": True, "message": "圆弧已绘制", "entity": arc}
            
        except Exception as e:
            err = _format_com_error(e)
            logger.error(f"绘制圆弧失败: {err}")
            return {"success": False, "message": f"绘制圆弧失败: {err}"}

    def draw_text(self, text: str, position: tuple, height: float = 2.5, layer: str = None) -> dict:
        """
        添加单行文字
        
        Args:
            text: 文字内容
            position: 插入点 (x, y) 或 (x, y, z)
            height: 字高
            layer: 图层名（可选）
        
        Returns:
            {"success": bool, "message": str, "entity": object}
        """
        if not self.ensure_document():
            return {"success": False, "message": "未连接到AutoCAD或无活动文档"}
        
        try:
            insert_point = self._make_point(position)
            text_obj = self.acad_doc.ModelSpace.AddText(text, insert_point, float(height))
            
            if layer:
                try:
                    text_obj.Layer = layer
                except Exception:
                    pass
            
            logger.info(f"已添加文字: '{text}'")
            return {"success": True, "message": f"文字已添加", "entity": text_obj}
            
        except Exception as e:
            err = _format_com_error(e)
            logger.error(f"添加文字失败: {err}")
            return {"success": False, "message": f"添加文字失败: {err}"}

    def draw_polyline(self, points: list, closed: bool = False, layer: str = None) -> dict:
        """
        绘制多段线
        
        Args:
            points: 点列表 [(x1,y1), (x2,y2), ...]
            closed: 是否闭合
            layer: 图层名（可选）
        
        Returns:
            {"success": bool, "message": str, "entity": object}
        """
        if not self.ensure_document():
            return {"success": False, "message": "未连接到AutoCAD或无活动文档"}
        
        if len(points) < 2:
            return {"success": False, "message": "多段线至少需要2个点"}
        
        try:
            point_array = self._make_point_2d(points)
            pline = self.acad_doc.ModelSpace.AddLightweightPolyline(point_array)
            pline.Closed = closed
            
            if layer:
                try:
                    pline.Layer = layer
                except Exception:
                    pass
            
            logger.info(f"已绘制多段线: {len(points)}个点")
            return {"success": True, "message": f"多段线已绘制（{len(points)}个点）", "entity": pline}
            
        except Exception as e:
            err = _format_com_error(e)
            logger.error(f"绘制多段线失败: {err}")
            return {"success": False, "message": f"绘制多段线失败: {err}"}

    def draw_polygon(self, center: tuple, radius: float, sides: int, layer: str = None) -> dict:
        """
        绘制正多边形（通过命令行，因为 COM API 没有直接方法）
        
        Args:
            center: 中心点 (x, y)
            radius: 外接圆半径
            sides: 边数（3-1024）
            layer: 图层名（可选）
        
        Returns:
            {"success": bool, "message": str}
        """
        if not self.ensure_document():
            return {"success": False, "message": "未连接到AutoCAD或无活动文档"}
        
        if sides < 3 or sides > 1024:
            return {"success": False, "message": "边数必须在3-1024之间"}
        
        try:
            x, y = float(center[0]), float(center[1])
            # 使用 POLYGON 命令
            cmd = f"_POLYGON {sides} {x},{y} I {radius}\n"
            self.acad_doc.SendCommand(cmd)
            
            logger.info(f"已绘制正{sides}边形")
            return {"success": True, "message": f"正{sides}边形已绘制"}
            
        except Exception as e:
            err = _format_com_error(e)
            logger.error(f"绘制多边形失败: {err}")
            return {"success": False, "message": f"绘制多边形失败: {err}"}

    def draw_star(
        self,
        center: tuple,
        outer_radius: float,
        inner_radius: float,
        points: int = 5,
        start_angle: float = 0.0,
        layer: str = None,
    ) -> dict:
        """
        绘制星形（用多段线 polyline 生成交替外/内半径顶点）。
        
        start_angle 使用弧度(rad)，points 表示星的尖数（例如国旗五角星=5）。
        """
        if not self.ensure_document():
            return {"success": False, "message": "未连接到AutoCAD或无活动文档"}

        try:
            if points < 3:
                return {"success": False, "message": "星形 points 必须 >= 3"}

            # 只使用 x/y 做 2D 星形；z 由 _make_point 决定
            cx = float(center[0]) if len(center) >= 1 else 0.0
            cy = float(center[1]) if len(center) >= 2 else 0.0
            r_out = float(outer_radius)
            r_in = float(inner_radius)
            a0 = float(start_angle)

            # 2*points 个顶点：外-内-外-内...
            # 每个相邻顶点夹角为 pi / points
            step = math.pi / float(points)
            pts = []
            for k in range(points * 2):
                ang = a0 + k * step
                r = r_out if (k % 2 == 0) else r_in
                x = cx + r * math.cos(ang)
                y = cy + r * math.sin(ang)
                pts.append((x, y))

            # 绘制 closed polyline 来表示星形
            point_array = self._make_point_2d(pts)
            pline = self.acad_doc.ModelSpace.AddLightweightPolyline(point_array)
            pline.Closed = True

            if layer:
                try:
                    pline.Layer = layer
                except Exception:
                    pass

            logger.info(f"已绘制星形: points={points}, outer={outer_radius}, inner={inner_radius}")
            return {"success": True, "message": "星形已绘制", "entity": pline}

        except Exception as e:
            err = _format_com_error(e)
            logger.error(f"绘制星形失败: {err}")
            return {"success": False, "message": f"绘制星形失败: {err}"}

    def set_layer(self, layer_name: str, create_if_not_exists: bool = True) -> dict:
        """
        设置当前图层
        
        Args:
            layer_name: 图层名
            create_if_not_exists: 不存在时是否创建
        
        Returns:
            {"success": bool, "message": str}
        """
        if not self.ensure_document():
            return {"success": False, "message": "未连接到AutoCAD或无活动文档"}
        
        try:
            layers = self.acad_doc.Layers
            
            # 检查图层是否存在
            layer = None
            try:
                layer = layers.Item(layer_name)
            except Exception:
                # 图层不存在
                if create_if_not_exists:
                    layer = layers.Add(layer_name)
                    logger.info(f"已创建图层: {layer_name}")
                else:
                    return {"success": False, "message": f"图层 '{layer_name}' 不存在"}
            
            if layer:
                self.acad_doc.ActiveLayer = layer
                logger.info(f"已设置当前图层: {layer_name}")
                return {"success": True, "message": f"当前图层已设置为 '{layer_name}'"}
            
        except Exception as e:
            err = _format_com_error(e)
            logger.error(f"设置图层失败: {err}")
            return {"success": False, "message": f"设置图层失败: {err}"}

    def zoom_extents(self) -> bool:
        """缩放到全部图形"""
        if not self.acad_app:
            return False
        try:
            self.acad_app.ZoomExtents()
            logger.info("已缩放到全部图形")
            return True
        except Exception as e:
            logger.warning(f"缩放失败: {_format_com_error(e)}")
            return False

    def zoom_center(self, center: tuple, magnification: float = 1.0) -> bool:
        """缩放到指定中心点"""
        if not self.acad_app:
            return False
        try:
            center_point = self._make_point(center)
            self.acad_app.ZoomCenter(center_point, float(magnification))
            return True
        except Exception as e:
            logger.warning(f"缩放失败: {_format_com_error(e)}")
            return False

    def get_entity_count(self) -> int:
        """获取模型空间中的实体数量"""
        if not self.ensure_document():
            return 0
        try:
            return self.acad_doc.ModelSpace.Count
        except Exception:
            return 0

    def delete_last_entity(self) -> dict:
        """删除最后一个绘制的实体"""
        if not self.ensure_document():
            return {"success": False, "message": "未连接到AutoCAD或无活动文档"}
        
        try:
            count = self.acad_doc.ModelSpace.Count
            if count > 0:
                last_entity = self.acad_doc.ModelSpace.Item(count - 1)
                last_entity.Delete()
                logger.info("已删除最后一个实体")
                return {"success": True, "message": "已删除最后一个实体"}
            else:
                return {"success": False, "message": "模型空间中没有实体"}
                
        except Exception as e:
            err = _format_com_error(e)
            logger.error(f"删除实体失败: {err}")
            return {"success": False, "message": f"删除失败: {err}"}

    def execute_drawing_commands(self, commands: list) -> dict:
        """
        批量执行绘图命令
        
        Args:
            commands: 绘图命令列表，格式如:
                [
                    {"type": "line", "start": [0, 0], "end": [100, 100]},
                    {"type": "circle", "center": [50, 50], "radius": 25},
                    {"type": "rectangle", "corner1": [0, 0], "corner2": [100, 50]}
                ]
        
        Returns:
            {"success": bool, "message": str, "results": list, "failed_count": int}
        """
        if not commands:
            return {"success": False, "message": "没有绘图命令", "results": [], "failed_count": 0}
        
        results = []
        failed_count = 0
        
        for cmd in commands:
            cmd_type = cmd.get("type", "").lower()
            result = {"type": cmd_type, "success": False}
            
            try:
                if cmd_type == "line":
                    result = self.draw_line(
                        cmd.get("start", (0, 0)),
                        cmd.get("end", (0, 0)),
                        cmd.get("layer")
                    )
                    result["type"] = "line"
                    
                elif cmd_type == "circle":
                    # 支持直径和半径两种参数
                    center = cmd.get("center", (0, 0))
                    radius = cmd.get("radius")
                    if radius is None:
                        diameter = cmd.get("diameter")
                        if diameter is not None:
                            radius = float(diameter) / 2
                        else:
                            radius = 10  # 默认半径
                    result = self.draw_circle(
                        center,
                        float(radius),
                        cmd.get("layer")
                    )
                    result["type"] = "circle"
                    
                elif cmd_type == "rectangle" or cmd_type == "rect":
                    result = self.draw_rectangle(
                        cmd.get("corner1", (0, 0)),
                        cmd.get("corner2", (100, 100)),
                        cmd.get("layer")
                    )
                    result["type"] = "rectangle"
                    
                elif cmd_type == "arc":
                    result = self.draw_arc(
                        cmd.get("center", (0, 0)),
                        cmd.get("radius", 10),
                        cmd.get("start_angle", 0),
                        cmd.get("end_angle", 3.14159),
                        cmd.get("layer")
                    )
                    result["type"] = "arc"
                    
                elif cmd_type == "text":
                    result = self.draw_text(
                        cmd.get("content", ""),
                        cmd.get("position", (0, 0)),
                        cmd.get("height", 2.5),
                        cmd.get("layer")
                    )
                    result["type"] = "text"
                    
                elif cmd_type == "polyline" or cmd_type == "pline":
                    result = self.draw_polyline(
                        cmd.get("points", []),
                        cmd.get("closed", False),
                        cmd.get("layer")
                    )
                    result["type"] = "polyline"
                    
                elif cmd_type == "polygon":
                    result = self.draw_polygon(
                        cmd.get("center", (0, 0)),
                        cmd.get("radius", 10),
                        cmd.get("sides", 6),
                        cmd.get("layer")
                    )
                    result["type"] = "polygon"
                    
                elif cmd_type == "star":
                    result = self.draw_star(
                        cmd.get("center", (0, 0)),
                        cmd.get("outer_radius", 10),
                        cmd.get("inner_radius", 5),
                        cmd.get("points", 5),
                        cmd.get("start_angle", 0.0),
                        cmd.get("layer"),
                    )
                    result["type"] = "star"
                    
                else:
                    result = {"success": False, "message": f"未知的绘图类型: {cmd_type}", "type": cmd_type}
                
            except Exception as e:
                result = {"success": False, "message": f"执行异常: {str(e)}", "type": cmd_type}
            
            if not result.get("success"):
                failed_count += 1
            results.append(result)
        
        success = failed_count == 0
        message = f"完成 {len(commands)} 个绘图命令，失败 {failed_count} 个"
        logger.info(message)
        
        return {
            "success": success,
            "message": message,
            "results": results,
            "failed_count": failed_count
        }

    # ==================== 图纸信息读取与Excel导出 ====================

    def get_layers_info(self) -> dict:
        """获取所有图层信息"""
        if not self.ensure_document():
            return {"success": False, "message": "未连接到AutoCAD或无活动文档", "layers": []}
        
        try:
            layers = self.acad_doc.Layers
            layer_list = []
            
            for i in range(layers.Count):
                layer = layers.Item(i)
                layer_info = {
                    "name": layer.Name,
                    "on": layer.LayerOn,  # 是否打开
                    "frozen": layer.Freeze,  # 是否冻结
                    "locked": layer.Lock,  # 是否锁定
                    "color": layer.Color,  # 颜色索引
                }
                layer_list.append(layer_info)
            
            logger.info(f"已获取 {len(layer_list)} 个图层信息")
            return {
                "success": True,
                "message": f"共 {len(layer_list)} 个图层",
                "layers": layer_list
            }
            
        except Exception as e:
            err = _format_com_error(e)
            logger.error(f"获取图层信息失败: {err}")
            return {"success": False, "message": f"获取失败: {err}", "layers": []}

    def _get_entity_geometry(self, entity) -> dict:
        """获取实体的几何详细信息（完整版，支持图形重建）"""
        geometry = {}
        entity_type = entity.ObjectName
        draw_cmd = None  # 绘制命令，供其他智能体使用
        
        try:
            if entity_type == "AcDbLine":
                # 直线：起点、终点、长度、角度
                start = entity.StartPoint
                end = entity.EndPoint
                geometry["start_x"] = round(start[0], 2)
                geometry["start_y"] = round(start[1], 2)
                geometry["start_z"] = round(start[2], 2)
                geometry["end_x"] = round(end[0], 2)
                geometry["end_y"] = round(end[1], 2)
                geometry["end_z"] = round(end[2], 2)
                geometry["length"] = round(entity.Length, 2)
                geometry["angle"] = round(entity.Angle, 2)
                # 绘制命令
                draw_cmd = {
                    "type": "line",
                    "start": [round(start[0], 2), round(start[1], 2)],
                    "end": [round(end[0], 2), round(end[1], 2)]
                }
                
            elif entity_type == "AcDbCircle":
                # 圆：圆心、半径、直径、面积、周长
                center = entity.Center
                radius = entity.Radius
                geometry["center_x"] = round(center[0], 2)
                geometry["center_y"] = round(center[1], 2)
                geometry["center_z"] = round(center[2], 2)
                geometry["radius"] = round(radius, 2)
                geometry["diameter"] = round(radius * 2, 2)
                geometry["area"] = round(3.14159 * radius * radius, 2)
                geometry["circumference"] = round(2 * 3.14159 * radius, 2)
                # 绘制命令
                draw_cmd = {
                    "type": "circle",
                    "center": [round(center[0], 2), round(center[1], 2)],
                    "radius": round(radius, 2)
                }
                
            elif entity_type == "AcDbArc":
                # 圆弧：圆心、半径、起始角度、终止角度、弧长
                center = entity.Center
                radius = entity.Radius
                start_angle = entity.StartAngle
                end_angle = entity.EndAngle
                geometry["center_x"] = round(center[0], 2)
                geometry["center_y"] = round(center[1], 2)
                geometry["radius"] = round(radius, 2)
                geometry["start_angle_deg"] = round(start_angle * 180 / 3.14159, 2)  # 角度制
                geometry["end_angle_deg"] = round(end_angle * 180 / 3.14159, 2)  # 角度制
                geometry["start_angle_rad"] = round(start_angle, 4)  # 弧度制
                geometry["end_angle_rad"] = round(end_angle, 4)  # 弧度制
                arc_length = radius * abs(end_angle - start_angle)
                geometry["arc_length"] = round(arc_length, 2)
                # 绘制命令
                draw_cmd = {
                    "type": "arc",
                    "center": [round(center[0], 2), round(center[1], 2)],
                    "radius": round(radius, 2),
                    "start_angle": round(start_angle, 4),
                    "end_angle": round(end_angle, 4)
                }
                
            elif entity_type in ("AcDbPolyline", "AcDb2dPolyline", "AcDb3dPolyline"):
                # 多段线：提取所有顶点坐标（关键！）
                try:
                    num_vertices = entity.NumberOfVertices
                    geometry["vertices_count"] = num_vertices
                except:
                    num_vertices = 0
                    geometry["vertices_count"] = "N/A"
                
                try:
                    geometry["length"] = round(entity.Length, 2)
                except:
                    geometry["length"] = "N/A"
                
                try:
                    geometry["closed"] = entity.Closed
                except:
                    geometry["closed"] = False
                
                try:
                    if entity.Closed:
                        geometry["area"] = round(entity.Area, 2)
                except:
                    pass
                
                # 提取所有顶点坐标（用于重建图形）
                vertices_list = []
                try:
                    for i in range(num_vertices):
                        coord = entity.Coordinate(i)
                        vertices_list.append([round(coord[0], 2), round(coord[1], 2)])
                    geometry["vertices_list"] = vertices_list
                    # 绘制命令
                    draw_cmd = {
                        "type": "polyline",
                        "vertices": vertices_list,
                        "closed": geometry.get("closed", False)
                    }
                except Exception as e:
                    geometry["vertices_error"] = str(e)
                    
            elif entity_type == "AcDbEllipse":
                # 椭圆：圆心、主轴半径、次轴半径
                center = entity.Center
                geometry["center_x"] = round(center[0], 2)
                geometry["center_y"] = round(center[1], 2)
                geometry["center_z"] = round(center[2], 2)
                try:
                    geometry["major_radius"] = round(entity.MajorRadius, 2)
                    geometry["minor_radius"] = round(entity.MinorRadius, 2)
                    # 绘制命令
                    draw_cmd = {
                        "type": "ellipse",
                        "center": [round(center[0], 2), round(center[1], 2)],
                        "major_radius": round(entity.MajorRadius, 2),
                        "minor_radius": round(entity.MinorRadius, 2)
                    }
                except:
                    pass
                    
            elif entity_type == "AcDbText":
                # 单行文字：位置、内容、高度、旋转角度
                pos = entity.InsertionPoint
                geometry["position_x"] = round(pos[0], 2)
                geometry["position_y"] = round(pos[1], 2)
                geometry["position_z"] = round(pos[2], 2)
                geometry["text"] = entity.TextString
                geometry["height"] = round(entity.Height, 2)
                try:
                    geometry["rotation"] = round(entity.Rotation, 2)
                except:
                    pass
                # 绘制命令
                draw_cmd = {
                    "type": "text",
                    "position": [round(pos[0], 2), round(pos[1], 2)],
                    "text": entity.TextString,
                    "height": round(entity.Height, 2)
                }
                
            elif entity_type == "AcDbMText":
                # 多行文字：位置、内容、高度、宽度
                pos = entity.InsertionPoint
                geometry["position_x"] = round(pos[0], 2)
                geometry["position_y"] = round(pos[1], 2)
                geometry["text"] = entity.TextString
                geometry["height"] = round(entity.Height, 2)
                try:
                    geometry["width"] = round(entity.Width, 2)
                except:
                    pass
                # 绘制命令
                draw_cmd = {
                    "type": "mtext",
                    "position": [round(pos[0], 2), round(pos[1], 2)],
                    "text": entity.TextString,
                    "height": round(entity.Height, 2)
                }
                
            elif entity_type == "AcDbPoint":
                # 点：坐标
                pos = entity.Coordinates
                geometry["x"] = round(pos[0], 2)
                geometry["y"] = round(pos[1], 2)
                geometry["z"] = round(pos[2], 2) if len(pos) > 2 else 0
                # 绘制命令
                draw_cmd = {
                    "type": "point",
                    "position": [round(pos[0], 2), round(pos[1], 2)]
                }
                
            elif entity_type in ("AcDbSpline", "AcDb2dSpline"):
                # 样条曲线：控制点坐标
                try:
                    num_ctrl = entity.NumberOfControlPoints
                    geometry["control_points_count"] = num_ctrl
                    geometry["degree"] = entity.Degree
                    # 提取控制点坐标
                    ctrl_points = []
                    for i in range(num_ctrl):
                        pt = entity.GetControlPoint(i)
                        ctrl_points.append([round(pt[0], 2), round(pt[1], 2)])
                    geometry["control_points"] = ctrl_points
                    # 绘制命令
                    draw_cmd = {
                        "type": "spline",
                        "control_points": ctrl_points,
                        "degree": entity.Degree
                    }
                except Exception as e:
                    geometry["spline_error"] = str(e)
                    
            elif entity_type == "AcDbHatch":
                # 填充：图案名、面积
                try:
                    geometry["pattern"] = entity.PatternName
                    geometry["area"] = round(entity.Area, 2)
                except:
                    pass
                    
            elif entity_type in ("AcDbSolid", "AcDbTrace"):
                # 实体填充：顶点坐标
                try:
                    coords = entity.Coordinates
                    points = []
                    for i in range(0, len(coords), 3):
                        points.append([round(coords[i], 2), round(coords[i+1], 2)])
                    geometry["points"] = points
                    geometry["area"] = round(entity.Area, 2)
                    # 绘制命令
                    draw_cmd = {
                        "type": "solid",
                        "points": points
                    }
                except:
                    pass
                    
            elif entity_type == "AcDbBlockReference":
                # 块引用：插入点、名称、缩放、旋转
                try:
                    pos = entity.InsertionPoint
                    geometry["insert_x"] = round(pos[0], 2)
                    geometry["insert_y"] = round(pos[1], 2)
                    geometry["insert_z"] = round(pos[2], 2)
                    geometry["block_name"] = entity.Name
                    geometry["rotation"] = round(entity.Rotation, 2)
                    try:
                        geometry["scale_x"] = round(entity.XScaleFactor, 2)
                        geometry["scale_y"] = round(entity.YScaleFactor, 2)
                        geometry["scale_z"] = round(entity.ZScaleFactor, 2)
                    except:
                        pass
                    # 绘制命令
                    draw_cmd = {
                        "type": "block",
                        "name": entity.Name,
                        "insert": [round(pos[0], 2), round(pos[1], 2)],
                        "rotation": round(entity.Rotation, 2)
                    }
                except:
                    pass
                    
            elif entity_type in ("AcDbRotatedDimension", "AcDbAlignedDimension", "AcDbDiametricDimension", "AcDbRadialDimension"):
                # 标注：测量值、位置
                try:
                    geometry["measurement"] = round(entity.Measurement, 2)
                    geometry["text"] = entity.TextOverride if entity.TextOverride else str(round(entity.Measurement, 2))
                except:
                    pass
                    
            # 通用属性
            try:
                geometry["color"] = entity.Color
            except:
                pass
            try:
                geometry["linetype"] = entity.Linetype
            except:
                pass
            try:
                geometry["lineweight"] = entity.Lineweight
            except:
                pass
            try:
                geometry["layer"] = entity.Layer
            except:
                pass
                
            # 添加绘制命令
            if draw_cmd:
                geometry["draw_command"] = draw_cmd
                
        except Exception as e:
            geometry["error"] = str(e)
            
        return geometry

    def get_entities_info(self) -> dict:
        """获取所有实体信息（含详细几何属性）"""
        if not self.ensure_document():
            return {"success": False, "message": "未连接到AutoCAD或无活动文档", "entities": []}
        
        try:
            model_space = self.acad_doc.ModelSpace
            entity_stats = {}
            entity_details = []
            
            for i in range(model_space.Count):
                entity = model_space.Item(i)
                entity_type = entity.ObjectName  # 如 AcDbLine, AcDbCircle
                
                # 统计数量
                if entity_type not in entity_stats:
                    entity_stats[entity_type] = 0
                entity_stats[entity_type] += 1
                
                # 获取实体基本信息
                try:
                    layer_name = entity.Layer
                except:
                    layer_name = "未知"
                
                # 获取几何详细信息
                geometry = self._get_entity_geometry(entity)
                
                entity_details.append({
                    "index": i + 1,
                    "type": entity_type,
                    "layer": layer_name,
                    "handle": entity.Handle,
                    "geometry": geometry
                })
            
            # 转换为列表格式
            stats_list = [{"type": k, "count": v} for k, v in entity_stats.items()]
            
            logger.info(f"已获取 {model_space.Count} 个实体，共 {len(stats_list)} 种类型")
            return {
                "success": True,
                "message": f"共 {model_space.Count} 个实体",
                "total_count": model_space.Count,
                "type_stats": stats_list,
                "entities": entity_details
            }
            
        except Exception as e:
            err = _format_com_error(e)
            logger.error(f"获取实体信息失败: {err}")
            return {"success": False, "message": f"获取失败: {err}", "entities": []}

    def get_drawing_info(self) -> dict:
        """获取完整图纸信息（图层 + 实体）"""
        if not self.ensure_document():
            return {"success": False, "message": "未连接到AutoCAD或无活动文档"}
        
        try:
            # 获取图纸基本信息
            doc_name = self.acad_doc.Name
            
            # 获取图层信息
            layers_result = self.get_layers_info()
            layers = layers_result.get("layers", [])
            
            # 获取实体信息
            entities_result = self.get_entities_info()
            type_stats = entities_result.get("type_stats", [])
            entities = entities_result.get("entities", [])
            
            # 按图层分组统计
            layer_entity_count = {}
            for entity in entities:
                layer = entity["layer"]
                if layer not in layer_entity_count:
                    layer_entity_count[layer] = 0
                layer_entity_count[layer] += 1
            
            logger.info(f"已获取图纸 {doc_name} 的完整信息")
            return {
                "success": True,
                "message": f"图纸信息已获取",
                "document_name": doc_name,
                "layers": layers,
                "layer_count": len(layers),
                "type_stats": type_stats,
                "total_entities": len(entities),
                "layer_entity_count": layer_entity_count,
                "entities": entities  # 添加完整的实体数据
            }
            
        except Exception as e:
            err = _format_com_error(e)
            logger.error(f"获取图纸信息失败: {err}")
            return {"success": False, "message": f"获取失败: {err}"}

    def export_to_excel(self, filepath: str, info_type: str = "all") -> dict:
        """
        导出图纸信息到Excel
        
        Args:
            filepath: Excel文件路径
            info_type: 导出类型 - "all"(全部), "layers"(图层), "entities"(实体)
        
        Returns:
            {"success": bool, "message": str, "filepath": str}
        """
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, Alignment, PatternFill
        except ImportError:
            return {"success": False, "message": "请先安装 openpyxl: pip install openpyxl"}
        
        if not self.ensure_document():
            return {"success": False, "message": "未连接到AutoCAD或无活动文档"}
        
        try:
            # 获取图纸信息
            drawing_info = self.get_drawing_info()
            if not drawing_info.get("success"):
                return {"success": False, "message": drawing_info.get("message", "获取图纸信息失败")}
            
            # 创建Excel工作簿
            wb = Workbook()
            
            # 样式定义
            header_font = Font(bold=True, size=12)
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            header_font_white = Font(bold=True, size=12, color="FFFFFF")
            center_align = Alignment(horizontal='center', vertical='center')
            
            # 工作表1: 图纸概览
            if info_type in ("all", "overview"):
                ws_overview = wb.active
                ws_overview.title = "图纸概览"
                
                overview_data = [
                    ["图纸信息导出报告"],
                    [""],
                    ["图纸名称", drawing_info.get("document_name", "未知")],
                    ["图层总数", drawing_info.get("layer_count", 0)],
                    ["实体总数", drawing_info.get("total_entities", 0)],
                    [""],
                    ["实体类型统计"],
                ]
                
                for row in overview_data:
                    ws_overview.append(row)
                
                # 添加实体类型统计
                ws_overview.append(["类型", "数量"])
                for stat in drawing_info.get("type_stats", []):
                    ws_overview.append([stat["type"], stat["count"]])
                
                # 设置标题样式
                ws_overview['A1'].font = Font(bold=True, size=16)
            
            # 工作表2: 图层信息
            if info_type in ("all", "layers"):
                ws_layers = wb.create_sheet("图层信息")
                
                # 表头
                headers = ["图层名称", "状态", "冻结", "锁定", "颜色索引", "实体数量"]
                ws_layers.append(headers)
                
                # 设置表头样式
                for col_idx, header in enumerate(headers, 1):
                    cell = ws_layers.cell(row=1, column=col_idx)
                    cell.font = header_font_white
                    cell.fill = header_fill
                    cell.alignment = center_align
                
                # 添加数据
                layer_entity_count = drawing_info.get("layer_entity_count", {})
                for layer in drawing_info.get("layers", []):
                    layer_name = layer["name"]
                    ws_layers.append([
                        layer_name,
                        "打开" if layer["on"] else "关闭",
                        "是" if layer["frozen"] else "否",
                        "是" if layer["locked"] else "否",
                        layer["color"],
                        layer_entity_count.get(layer_name, 0)
                    ])
                
                # 调整列宽
                ws_layers.column_dimensions['A'].width = 30
                ws_layers.column_dimensions['B'].width = 10
                ws_layers.column_dimensions['C'].width = 10
                ws_layers.column_dimensions['D'].width = 10
                ws_layers.column_dimensions['E'].width = 12
                ws_layers.column_dimensions['F'].width = 12
            
            # 工作表3: 实体明细（含几何信息）
            if info_type in ("all", "entities"):
                ws_entities = wb.create_sheet("实体明细")
                
                # 获取所有实体
                entities = drawing_info.get("entities", [])
                
                # 收集所有可能的几何属性名
                all_geometry_keys = set()
                for entity in entities:
                    geometry = entity.get("geometry", {})
                    # 排除复杂对象（列表、字典）
                    for k, v in geometry.items():
                        if not isinstance(v, (list, dict)):
                            all_geometry_keys.add(k)
                
                # 排序几何属性，按优先级排列
                priority_keys = [
                    "center_x", "center_y", "center_z",
                    "start_x", "start_y", "start_z",
                    "end_x", "end_y", "end_z",
                    "position_x", "position_y", "position_z",
                    "insert_x", "insert_y", "insert_z",
                    "x", "y", "z",
                    "radius", "diameter",
                    "length", "width", "height",
                    "area", "circumference", "arc_length",
                    "angle", "rotation", "start_angle_deg", "end_angle_deg",
                    "major_radius", "minor_radius",
                    "vertices_count", "closed", "degree",
                    "text", "pattern", "block_name", "measurement",
                    "color", "linetype", "lineweight"
                ]
                # 过滤出实际存在的属性
                sorted_geometry_keys = [k for k in priority_keys if k in all_geometry_keys]
                # 添加其他未列出的属性
                for k in sorted(all_geometry_keys - set(priority_keys)):
                    sorted_geometry_keys.append(k)
                
                # 动态表头：基本信息 + 几何属性 + 顶点坐标 + 绘制命令
                base_headers = ["序号", "实体类型", "所属图层", "句柄"]
                geometry_headers = [_format_geometry_key(k) for k in sorted_geometry_keys]
                extra_headers = ["顶点坐标列表", "绘制命令(JSON)"]
                headers = base_headers + geometry_headers + extra_headers
                ws_entities.append(headers)
                
                # 设置表头样式
                for col_idx, header in enumerate(headers, 1):
                    cell = ws_entities.cell(row=1, column=col_idx)
                    cell.font = header_font_white
                    cell.fill = header_fill
                    cell.alignment = center_align
                
                # 添加数据
                import json
                for idx, entity in enumerate(entities, 1):
                    geometry = entity.get("geometry", {})
                    row_data = [
                        idx,
                        entity["type"],
                        entity["layer"],
                        entity["handle"]
                    ]
                    # 添加几何属性值（排除复杂对象）
                    for key in sorted_geometry_keys:
                        value = geometry.get(key, "")
                        row_data.append(value)
                    
                    # 添加顶点坐标列表（转为字符串）
                    vertices_list = geometry.get("vertices_list", [])
                    if vertices_list:
                        vertices_str = json.dumps(vertices_list, ensure_ascii=False)
                    else:
                        vertices_str = ""
                    row_data.append(vertices_str)
                    
                    # 添加绘制命令（转为字符串）
                    draw_cmd = geometry.get("draw_command", {})
                    if draw_cmd:
                        draw_cmd_str = json.dumps(draw_cmd, ensure_ascii=False)
                    else:
                        draw_cmd_str = ""
                    row_data.append(draw_cmd_str)
                    
                    ws_entities.append(row_data)
                
                # 调整列宽
                ws_entities.column_dimensions['A'].width = 8   # 序号
                ws_entities.column_dimensions['B'].width = 18  # 实体类型
                ws_entities.column_dimensions['C'].width = 20  # 所属图层
                ws_entities.column_dimensions['D'].width = 12  # 句柄
                # 几何属性列宽
                col_idx = 5
                for key in sorted_geometry_keys:
                    col_letter = chr(64 + col_idx) if col_idx <= 26 else chr(64 + (col_idx - 1) // 26) + chr(65 + (col_idx - 1) % 26)
                    if "text" in key.lower():
                        ws_entities.column_dimensions[col_letter].width = 30
                    elif key in ("radius", "diameter", "length", "area"):
                        ws_entities.column_dimensions[col_letter].width = 12
                    else:
                        ws_entities.column_dimensions[col_letter].width = 15
                    col_idx += 1
                # 顶点坐标和绘制命令列宽
                for _ in range(2):
                    col_letter = chr(64 + col_idx) if col_idx <= 26 else chr(64 + (col_idx - 1) // 26) + chr(65 + (col_idx - 1) % 26)
                    ws_entities.column_dimensions[col_letter].width = 50
                    col_idx += 1
            
            # 保存文件
            wb.save(filepath)
            logger.info(f"已导出图纸信息到: {filepath}")
            
            return {
                "success": True,
                "message": f"已成功导出到: {filepath}",
                "filepath": filepath
            }
            
        except Exception as e:
            err = _format_com_error(e)
            logger.error(f"导出Excel失败: {err}")
            return {"success": False, "message": f"导出失败: {err}"}
